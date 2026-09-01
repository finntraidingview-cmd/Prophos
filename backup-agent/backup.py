#!/usr/bin/env python3
# ════════════════════════════════════════════════════════════════════════════
# Prophos Backup-Agent (01.09.2026)
#
# Zieht einmal täglich ALLE Tabellen aus Supabase (Projekt gxhkannmzpyuepxlepta)
# und legt sie als vollständigen Tages-Schnappschuss ab — Finns Versicherung,
# falls Supabase mal down ist oder Daten verloren gehen ("einen Agenten, der
# mir jeden Tag ein Backup macht").
#
# Warum VOLL-Schnappschuss statt "nur das Neue anhängen":
#   Ein kompletter Abzug pro Tag enthält automatisch alles Neue UND alle
#   Änderungen/Löschungen. Mit 30 Tagen Aufbewahrung gibt es zusätzlich eine
#   Zeitmaschine: versehentlich gelöschte Daten stehen im Schnappschuss von
#   gestern. Bei aktuell ~4.300 Zeilen über 29 Tabellen ist das eine Sache
#   von Sekunden — Anhängen wäre komplexer und fehleranfälliger für nichts.
#
# Ablage: ~/Prophos-Backups/JJJJ-MM-TT/  (PRIMÄR, lokal)
#   - eine CSV pro Tabelle (maschinenlesbar, für Wiederherstellung/Import)
#   - EINE Excel-Mappe mit allen Tabellen als Blätter (für Google Sheets:
#     einfach auf sheets.google.com reinziehen)
#   - _ANSEHEN.html — alle Tabellen im Browser, weil auf dem Mac KEIN
#     Tabellenprogramm installiert ist (festgestellt 01.09.2026: kein Numbers,
#     kein Excel — die xlsx ging schlicht nicht auf)
#   - _INFO.txt mit Zeilenzahlen und etwaigen Fehlern
#   Danach Spiegel-Versuch nach iCloud Drive/Prophos-Backups (überlebt einen
#   Mac-Ausfall). WICHTIG: macOS verweigert launchd-Hintergrundjobs den
#   iCloud-Zugriff (TCC, beobachtet 01.09.2026) — der Spiegel klappt also nur,
#   wenn python3 unter Systemeinstellungen → Datenschutz → Festplattenvollzugriff
#   freigeschaltet ist. Ohne das bleibt das Backup trotzdem vollständig, nur lokal.
#
# Zugang: Nur-Lese-Rolle backup_reader (sql/2026-09-01_backup-reader.sql),
# Zugangsdaten NUR lokal in ~/.prophos/backup-config.json (chmod 600).
# TLS voll verifiziert gegen die gepinnte Supabase-Root-CA.
#
# ABSICHTLICH laut scheitern: Eine neue Tabelle ohne GRANT/Policy erzeugt
# "permission denied" statt still leer gesichert zu werden. Dann einfach
# sql/2026-09-01_backup-reader.sql erneut einspielen (idempotent).
# ════════════════════════════════════════════════════════════════════════════
import csv
import datetime
import decimal
import html
import json
import os
import re
import shutil
import ssl
import sys
import uuid

# launchd startet ohne Login-Shell — die per "pip install --user" installierten
# Module (pg8000, openpyxl) liegen im User-Site-Verzeichnis, das hier zur
# Sicherheit explizit vorn eingehängt wird.
sys.path.insert(0, os.path.expanduser("~/Library/Python/%d.%d/lib/python/site-packages"
                                      % sys.version_info[:2]))
import pg8000.native
from openpyxl import Workbook

AUFBEWAHRUNG_TAGE = 30
KONFIG = os.path.expanduser("~/.prophos/backup-config.json")
LOKAL_BASIS = os.path.expanduser("~/Prophos-Backups")
ICLOUD_BASIS = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Prophos-Backups")


def log(msg):
    print(f"[{datetime.datetime.now():%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


def verbinden(cfg):
    """Probiert die Host-Kandidaten der Reihe nach (Pooler zuerst — der
    direkte DB-Host ist nur per IPv6 erreichbar, das kann sich aber ändern)."""
    ctx = ssl.create_default_context(cafile=os.path.expanduser(cfg["ca_file"]))
    letzter_fehler = None
    for kand in cfg["host_candidates"]:
        try:
            conn = pg8000.native.Connection(
                kand["user"], host=kand["host"], port=kand["port"],
                database=cfg["database"], password=cfg["password"],
                ssl_context=ctx, timeout=20)
            log(f"Verbunden über {kand['host']}")
            return conn
        except Exception as e:
            letzter_fehler = e
            log(f"Kandidat {kand['host']} nicht erreichbar: {e}")
    raise RuntimeError(f"Keine Verbindung zu Supabase möglich: {letzter_fehler}")


def als_text(wert):
    """Zellwert für CSV/Excel: JSON bleibt JSON, Datum wird ISO, Rest wird Text."""
    if wert is None:
        return ""
    if isinstance(wert, (dict, list)):
        return json.dumps(wert, ensure_ascii=False)
    if isinstance(wert, (datetime.datetime, datetime.date, datetime.time)):
        return wert.isoformat()
    if isinstance(wert, decimal.Decimal):
        return str(wert)
    if isinstance(wert, uuid.UUID):
        return str(wert)
    if isinstance(wert, (bytes, bytearray, memoryview)):
        return "\\x" + bytes(wert).hex()
    return wert  # str, int, float, bool direkt


def blattname(name, vergeben):
    """Excel erlaubt max. 31 Zeichen pro Blattname — kürzen und eindeutig halten."""
    kurz = name[:31]
    n = 2
    while kurz in vergeben:
        suffix = f"~{n}"
        kurz = name[:31 - len(suffix)] + suffix
        n += 1
    vergeben.add(kurz)
    return kurz


def alte_schnappschuesse_loeschen(basis):
    """Rotation: nur die letzten AUFBEWAHRUNG_TAGE Tagesordner behalten."""
    muster = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    ordner = sorted(d for d in os.listdir(basis)
                    if muster.match(d) and os.path.isdir(os.path.join(basis, d)))
    for alt in ordner[:-AUFBEWAHRUNG_TAGE]:
        shutil.rmtree(os.path.join(basis, alt))
        log(f"Alten Schnappschuss entfernt: {basis}/{alt}")


def nach_icloud_spiegeln(heute):
    """Best-effort-Kopie des Tagesordners nach iCloud Drive. Scheitert leise
    mit Warnung, wenn macOS dem launchd-Kontext den Zugriff verweigert."""
    try:
        if not os.path.isdir(os.path.dirname(ICLOUD_BASIS)):
            log("WARNUNG: iCloud Drive nicht vorhanden — kein Spiegel")
            return
        ziel = os.path.join(ICLOUD_BASIS, heute)
        shutil.copytree(os.path.join(LOKAL_BASIS, heute), ziel, dirs_exist_ok=True)
        alte_schnappschuesse_loeschen(ICLOUD_BASIS)
        log(f"Nach iCloud gespiegelt: {ziel}")
    except (PermissionError, shutil.Error):
        # shutil.Error von copytree enthält EINE Zeile pro Datei — nicht den
        # ganzen Roman loggen, die Ursache ist immer dieselbe (TCC)
        log("WARNUNG: iCloud-Spiegel verweigert (launchd hat keinen iCloud-Zugriff; "
            "Abhilfe: python3 unter Festplattenvollzugriff eintragen). "
            "Backup liegt vollständig lokal in " + LOKAL_BASIS)
    except Exception as e:
        log(f"WARNUNG: iCloud-Spiegel fehlgeschlagen: {e}")


def ansicht_schreiben(ziel, heute, daten, ergebnis):
    """_ANSEHEN.html: alle Tabellen als eine Browser-Seite. Der Mac hat kein
    Tabellenprogramm (kein Numbers/Excel, 01.09.2026) — das hier öffnet überall."""
    esc = lambda v: html.escape(str(v))
    teile = ["""<!doctype html><html lang="de"><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Prophos-Backup """ + heute + """</title>
<style>
  body { font: 14px -apple-system, sans-serif; margin: 0; background: #f5f6f8; color: #1c1e21; }
  header { padding: 18px 24px 6px; }
  h1 { font-size: 19px; margin: 0 0 4px; }
  .meta { color: #667; font-size: 13px; }
  nav { padding: 8px 24px 14px; display: flex; flex-wrap: wrap; gap: 6px; }
  nav a { background: #fff; border: 1px solid #d5d9e0; border-radius: 14px;
          padding: 3px 10px; text-decoration: none; color: #234; font-size: 12.5px; }
  nav a span { color: #889; }
  section { margin: 0 24px 26px; }
  h2 { font-size: 15px; margin: 0 0 6px; }
  h2 span { color: #889; font-weight: normal; font-size: 13px; }
  .wrap { overflow-x: auto; background: #fff; border: 1px solid #d5d9e0; border-radius: 8px; }
  table { border-collapse: collapse; font-size: 12.5px; white-space: nowrap; }
  th, td { padding: 4px 10px; border-bottom: 1px solid #eceff3; text-align: left;
           max-width: 340px; overflow: hidden; text-overflow: ellipsis; }
  th { position: sticky; top: 0; background: #f0f2f5; font-weight: 600; }
  tr:hover td { background: #f7fafd; }
</style>
<header><h1>Prophos-Backup """ + heute + "</h1><div class='meta'>"
             + esc(f"{len(daten)} Tabellen · {sum(len(z) for _, z in daten.values())} Zeilen"
                   + " · vollständige Daten in den CSVs daneben") + "</div></header><nav>"]
    for tab, z, _ in ergebnis:
        if tab in daten:
            teile.append(f"<a href='#{esc(tab)}'>{esc(tab)} <span>{z}</span></a>")
    teile.append("</nav>")
    for tab, (spalten, zeilen) in daten.items():
        teile.append(f"<section><h2 id='{esc(tab)}'>{esc(tab)} <span>{len(zeilen)} Zeilen</span></h2>"
                     "<div class='wrap'><table><tr>"
                     + "".join(f"<th>{esc(s)}</th>" for s in spalten) + "</tr>")
        for z in zeilen:
            teile.append("<tr>" + "".join(f"<td>{esc(v)}</td>" for v in z) + "</tr>")
        teile.append("</table></div></section>")
    with open(os.path.join(ziel, "_ANSEHEN.html"), "w", encoding="utf-8") as f:
        f.write("".join(teile))


def main():
    cfg = json.load(open(KONFIG))

    heute = f"{datetime.date.today():%Y-%m-%d}"
    ziel = os.path.join(LOKAL_BASIS, heute)
    os.makedirs(ziel, exist_ok=True)
    os.chmod(LOKAL_BASIS, 0o700)   # Backups (inkl. Zugangsdaten-Tabellen) nur für Finn

    conn = verbinden(cfg)
    tabellen = [r[0] for r in conn.run(
        "select tablename from pg_tables where schemaname = 'public' order by tablename")]
    log(f"{len(tabellen)} Tabellen gefunden")

    mappe = Workbook()
    mappe.remove(mappe.active)  # openpyxl legt sonst ein leeres "Sheet" an
    vergebene_blaetter = set()
    daten = {}      # tabelle -> (spalten, textzeilen) für die Browser-Ansicht
    ergebnis = []   # (tabelle, zeilen | None, fehlertext)
    fehler = 0

    for tab in tabellen:
        try:
            try:
                # ORDER BY 1 macht die Ausgabe stabil (Tages-Diffs vergleichbar);
                # falls die erste Spalte mal nicht sortierbar ist: unsortiert
                zeilen = conn.run(f'select * from public."{tab}" order by 1')
            except Exception:
                zeilen = conn.run(f'select * from public."{tab}"')
            spalten = [c["name"] for c in conn.columns]
            textzeilen = [[als_text(v) for v in z] for z in zeilen]

            with open(os.path.join(ziel, f"{tab}.csv"), "w", newline="",
                      encoding="utf-8-sig") as f:   # BOM: Excel zeigt Umlaute korrekt
                w = csv.writer(f)
                w.writerow(spalten)
                w.writerows(textzeilen)

            blatt = mappe.create_sheet(blattname(tab, vergebene_blaetter))
            blatt.append(spalten)
            for z in textzeilen:
                blatt.append(z)
            blatt.freeze_panes = "A2"

            daten[tab] = (spalten, textzeilen)
            ergebnis.append((tab, len(textzeilen), ""))
        except Exception as e:
            fehler += 1
            ergebnis.append((tab, None, str(e)))
            log(f"FEHLER bei Tabelle {tab}: {e}")

    conn.close()

    if not mappe.sheetnames:   # alle Tabellen gescheitert — leere Mappe wäre invalide
        mappe.create_sheet("leer")
    mappe.save(os.path.join(ziel, f"Prophos-Backup_{heute}.xlsx"))

    ansicht_schreiben(ziel, heute, daten, ergebnis)

    gesamt = sum(z for _, z, _ in ergebnis if z is not None)
    with open(os.path.join(ziel, "_INFO.txt"), "w", encoding="utf-8") as f:
        f.write(f"Prophos-Backup vom {heute}\n")
        f.write(f"Quelle: Supabase-Projekt gxhkannmzpyuepxlepta (Nur-Lese-Rolle backup_reader)\n")
        f.write(f"{len(tabellen)} Tabellen, {gesamt} Zeilen, {fehler} Fehler\n\n")
        for tab, z, err in ergebnis:
            f.write(f"  {tab:45s} {'FEHLER: ' + err if z is None else str(z) + ' Zeilen'}\n")
        f.write("\nWiederherstellung: CSVs pro Tabelle importieren (Supabase: Table Editor\n"
                "→ Import data from CSV) oder die Excel-Mappe als Nachschlagewerk nutzen.\n"
                "Bei 'permission denied'-Fehlern: sql/2026-09-01_backup-reader.sql erneut\n"
                "einspielen (neue Tabelle ohne Leserecht).\n")

    alte_schnappschuesse_loeschen(LOKAL_BASIS)
    log(f"Fertig: {len(tabellen) - fehler}/{len(tabellen)} Tabellen, "
        f"{gesamt} Zeilen → {ziel}")
    nach_icloud_spiegeln(heute)
    if fehler:
        log(f"ACHTUNG: {fehler} Tabelle(n) NICHT gesichert — Details in _INFO.txt")
        sys.exit(1)


if __name__ == "__main__":
    main()
